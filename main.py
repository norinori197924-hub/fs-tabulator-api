#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
for survey 集計API サーバー
FastAPI + uvicorn で動作
"""

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import tempfile, os, shutil, csv, datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = FastAPI(title="for survey 集計API")

# CORS設定（WADAXのサイトからアクセスできるようにする）
app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://www.pileup-datalib.com", "http://localhost", "http://127.0.0.1"],   # 本番環境では "https://www.pileup-datalib.com" に変更
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

ENCODING = 'shift-jis'
DECIMAL  = 2
ZERO_DASH = True

# ============================================================
# カラー定義
# ============================================================
C = {
    'h_navy'  : '1F3864', 'h_fg'    : 'FFFFFF',
    'q_hd'    : '2E5DA6', 'q_fg'    : 'FFFFFF',
    'n_row'   : 'D6E4F0', 'n_fg'    : '1F3864',
    'row_odd' : 'FFFFFF', 'row_even': 'F2F7FB',
    'total_bg': 'EBF3FB', 'hi_green': 'E2F0D9',
    'gray'    : 'F4F4F4', 'border'  : 'BDD0E4',
    'toc_odd' : 'F0F5FB', 'toc_even': 'FFFFFF',
    'link_fg' : '1155CC', 'back_bg' : 'E8F0FE',
    'wb_bg'   : 'FFF3CC',
}

def hf(h): return PatternFill('solid', start_color=h, fgColor=h)
def al(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin(color=None):
    c = color or C['border']
    s = Side(style='thin', color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def medium(color=None):
    c = color or C['h_navy']
    s = Side(style='medium', color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def ac(ws, row, col, val, bold=False, bg=None, fg='000000',
       h='left', v='center', sz=9, wrap=False, bdr=None):
    c = ws.cell(row, col, val)
    c.font = Font(bold=bold, color=fg, size=sz, name='Arial')
    if bg: c.fill = hf(bg)
    c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    c.border = bdr or thin()
    return c
def fmt_pct(v):
    if ZERO_DASH and v == 0: return '—'
    return f'{v:.{DECIMAL}f}%'

# ============================================================
# パーサー
# ============================================================
def parse_layout(text):
    rows = list(csv.reader(text.splitlines(), delimiter='\t'))
    meta = {}
    for i, c in enumerate(rows[0]):
        c = c.strip()
        if c in ('調査名','調査ID','有効サンプル数') and i+1 < len(rows[0]):
            meta[c] = rows[0][i+1].strip()
    questions, cur = {}, None
    for row in rows[2:]:
        g = lambda i: row[i].strip() if i < len(row) else ''
        qn,qt_c,c2,c3,ans,col,cho,text = g(0),g(1),g(2),g(3),g(4),g(6),g(7),g(8)
        if qn and not cho:
            cur = qn
            if qn not in questions:
                questions[qn] = {'q_type':qt_c,'ans_type':ans,'label':text,
                                  'col':col,'items':{},'sub_items':[]}
            continue
        if not qn and not c2 and not c3 and cho and cur:
            questions[cur]['items'][cho] = text; continue
        if not qn and col and cur:
            nm = c3 if c3 else c2
            if nm:
                questions[cur]['sub_items'].append(
                    {'item':nm,'col':col,'choice_no':cho,'text':text})
    return meta, questions

def parse_rawdata(text):
    rows = list(csv.reader(text.splitlines(), delimiter='\t'))
    headers = rows[0]
    hi = {h:i for i,h in enumerate(headers)}
    sta = hi.get('STA')
    data = [r for r in rows[1:]
            if sta is not None and len(r) > sta and r[sta] == 'COMP']
    return headers, hi, data

def detect_type(q):
    ans = q.get('ans_type','').upper(); qt = q.get('q_type','').upper()
    if 'SA' in ans: return 'SA'
    if 'MA' in ans: return 'MA'
    if 'FA' in ans: return 'FA'
    if 'NUM' in ans: return 'NUM'
    if qt == 'M': return 'MA'
    if qt == 'S': return 'SA'
    if qt in ('F','FS'): return 'FA'
    return 'SA'

# ============================================================
# 集計エンジン
# ============================================================
def tabulate_sa(qid, questions, data, hi):
    q = questions.get(qid)
    if not q: return None
    N = len(data)
    cnt = defaultdict(float)
    for row in data:
        v = (row[hi[qid]] if qid in hi and hi[qid] < len(row) else '').strip()
        cnt[v] += 1
    items = q.get('items', {})
    codes = sorted([k for k in cnt if k != ''],
                   key=lambda x: int(x) if x.lstrip('-').isdigit() else 999)
    return {'qid':qid,'qtype':'SA','label':q['label'],
            'codes':codes,'items':items,'cnt':cnt,'N':N}

def tabulate_ma(qid, questions, data, hi):
    q = questions.get(qid)
    if not q: return None
    N = len(data)
    subs = [s for s in q.get('sub_items',[]) if s['item'] in hi]
    rows_out = []
    for sub in subs:
        col = sub['item']
        n = sum(1 for row in data
                if (row[hi[col]] if hi[col] < len(row) else '') == '1')
        rows_out.append({'no':sub['choice_no'],'text':sub['text'],'n':n})
    return {'qid':qid,'qtype':'MA','label':q['label'],'rows':rows_out,'N':N}

def cross_engine(ax_qid, tg_qid, questions, data, hi):
    qax = questions.get(ax_qid); qtg = questions.get(tg_qid)
    if not qax or not qtg: return None
    N = len(data)
    ax_items = qax.get('items',{})
    ax_codes = sorted(ax_items, key=lambda x:(int(x) if x.lstrip('-').isdigit() else 999))
    ax_idx = defaultdict(list)
    for i, row in enumerate(data):
        v = (row[hi[ax_qid]] if ax_qid in hi and hi[ax_qid]<len(row) else '').strip()
        ax_idx[v].append(i)
    all_idx = list(range(N))
    ax_bases = {ac:len(ax_idx.get(ac,[])) for ac in ax_codes}
    tg_type = detect_type(qtg)
    rows_out = []
    if tg_type == 'SA':
        tg_items = qtg.get('items',{})
        tg_codes = sorted(tg_items, key=lambda x:(int(x) if x.lstrip('-').isdigit() else 999))
        for tc in tg_codes:
            def cnt_sa(idxs):
                return sum(1 for i in idxs
                           if (data[i][hi[tg_qid]] if hi[tg_qid]<len(data[i]) else '').strip()==tc)
            by_ax = {ac:{'n':cnt_sa(ax_idx.get(ac,[])),'base':ax_bases[ac]}
                     for ac in ax_codes}
            rows_out.append({'code':tc,'label':tg_items.get(tc,tc),
                             'n_all':cnt_sa(all_idx),'by_ax':by_ax})
    elif tg_type == 'MA':
        subs = [s for s in qtg.get('sub_items',[]) if s['item'] in hi]
        for sub in subs:
            col = sub['item']
            def cnt_ma(idxs):
                return sum(1 for i in idxs
                           if (data[i][hi[col]] if hi[col]<len(data[i]) else '')=='1')
            by_ax = {ac:{'n':cnt_ma(ax_idx.get(ac,[])),'base':ax_bases[ac]}
                     for ac in ax_codes}
            rows_out.append({'code':sub['choice_no'],'label':sub['text'],
                             'n_all':cnt_ma(all_idx),'by_ax':by_ax})
    return {'ax_qid':ax_qid,'tg_qid':tg_qid,
            'ax_label':qax['label'],'tg_label':qtg['label'],
            'ax_codes':ax_codes,'ax_items':ax_items,'ax_bases':ax_bases,
            'rows':rows_out,'N':N}

# ============================================================
# Excel生成（GT）
# ============================================================
def build_gt_excel(meta, questions, data, hi, target_qids, filepath):
    wb = Workbook(); wb.remove(wb.active)
    N = len(data); TODAY = datetime.date.today().strftime('%Y年%m月%d日')
    SNAME = 'GT集計'

    # --- 表紙 ---
    ws_toc = wb.create_sheet('表紙')
    ws_toc.sheet_view.showGridLines = False
    ws_toc.column_dimensions['A'].width = 3
    ws_toc.column_dimensions['B'].width = 14
    ws_toc.column_dimensions['C'].width = 18
    ws_toc.column_dimensions['D'].width = 46
    ws_toc.column_dimensions['E'].width = 10
    ws_toc.merge_cells('B1:E1'); ws_toc.row_dimensions[1].height = 44
    t = ws_toc['B1']; t.value = 'GT集計表'
    t.font = Font(bold=True,size=22,color='FFFFFF',name='Arial')
    t.fill = hf(C['h_navy']); t.alignment = al('center')
    infos = [('調査名',meta.get('調査名','—')),('調査ID',meta.get('調査ID','—')),
             ('有効回答数',f'{N:,} サンプル'),('出力日',TODAY)]
    for ri,(k,v) in enumerate(infos,3):
        ws_toc.row_dimensions[ri].height = 20
        kc = ws_toc.cell(ri,2,k)
        kc.font = Font(bold=True,color=C['h_fg']); kc.fill = hf(C['q_hd'])
        kc.alignment = al('center'); kc.border = thin(C['h_navy'])
        vc = ws_toc.cell(ri,3,v)
        vc.font = Font(size=9,name='Arial'); vc.fill = hf(C['gray'])
        vc.alignment = al(); vc.border = thin(C['h_navy'])
        ws_toc.merge_cells(start_row=ri,start_column=4,end_row=ri,end_column=5)
        ws_toc.cell(ri,4).border = thin(C['h_navy'])

    ws_toc.row_dimensions[8].height = 20
    for ci,lbl in enumerate(['No.','設問番号','タイプ','設問文','選択肢数'],2):
        c = ws_toc.cell(8,ci,lbl)
        c.font = Font(bold=True,color=C['h_fg'],size=9,name='Arial')
        c.fill = hf(C['h_navy']); c.alignment = al('center'); c.border = thin(C['h_navy'])

    # 行番号を事前計算
    cur_row = 3
    q_row_map = {}
    for qid in target_qids:
        q = questions.get(qid)
        if not q: continue
        qt = detect_type(q)
        if qt in ('FA','D'): continue
        q_row_map[qid] = cur_row
        if qt == 'SA':
            nc = len([k for k in q.get('items',{}) if k != ''])
        else:
            nc = len([s for s in q.get('sub_items',[]) if s['item'] in hi])
        cur_row += 1 + 1 + nc + 1

    toc_row = 9; no = 1
    for qid in target_qids:
        q = questions.get(qid)
        if not q: continue
        qt = detect_type(q)
        if qt in ('FA','D'): continue
        bg = C['toc_odd'] if no%2==0 else C['toc_even']
        dest_row = q_row_map.get(qid,3)
        ws_toc.row_dimensions[toc_row].height = 17
        ac(ws_toc,toc_row,2,no,bg=bg,h='center')
        lc = ws_toc.cell(toc_row,3); lc.value = qid
        lc.hyperlink = f"#{SNAME}!A{dest_row}"
        lc.font = Font(bold=True,color=C['link_fg'],underline='single',size=9,name='Arial')
        lc.fill = hf(bg); lc.alignment = al('center'); lc.border = thin()
        badge = {'SA':'SA（単択）','MA':'MA（複数）'}.get(qt,qt)
        ac(ws_toc,toc_row,4,badge,bg=bg,h='center')
        ac(ws_toc,toc_row,5,(q.get('label','') or '')[:40],bg=bg)
        if qt == 'SA': nc = len([k for k in q.get('items',{}) if k!=''])
        else: nc = len([s for s in q.get('sub_items',[]) if s['item'] in hi])
        ac(ws_toc,toc_row,6,nc,bg=bg,h='center')
        toc_row += 1; no += 1

    # --- GT集計シート ---
    ws = wb.create_sheet(SNAME); ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A3'
    ws.column_dimensions['A'].width = 12; ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 30; ws.column_dimensions['D'].width = 9
    ws.column_dimensions['E'].width = 10

    ws.row_dimensions[1].height = 22
    ws.merge_cells('A1:E1')
    t2 = ws.cell(1,1,'GT集計（単純集計）')
    t2.font = Font(bold=True,size=11,color='FFFFFF',name='Arial')
    t2.fill = hf(C['h_navy']); t2.alignment = al('center')

    ws.row_dimensions[2].height = 18
    for ci,h in enumerate(['設問番号','設問文','選択肢','実数 n','実数 %'],1):
        c = ws.cell(2,ci,h)
        c.font = Font(bold=True,size=9,color=C['h_fg'],name='Arial')
        c.fill = hf(C['q_hd']); c.alignment = al('center'); c.border = thin(C['h_navy'])

    row = 3
    for qid in target_qids:
        q = questions.get(qid)
        if not q: continue
        qt = detect_type(q)
        if qt in ('FA','D'): continue

        ws.row_dimensions[row].height = 28
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
        hc = ws.cell(row,1); hc.value = f'【{qid}】{q["label"]}'
        hc.font = Font(bold=True,size=9,color=C['q_fg'],name='Arial')
        hc.fill = hf(C['q_hd']); hc.alignment = al('left',wrap=True); hc.border = medium()
        bc = ws.cell(row,5); bc.value = '▲表紙へ'
        bc.hyperlink = '#表紙!A1'
        bc.font = Font(color=C['link_fg'],underline='single',size=8,italic=True,name='Arial')
        bc.fill = hf(C['back_bg']); bc.alignment = al('center'); bc.border = thin()
        row += 1

        ws.row_dimensions[row].height = 16
        ac(ws,row,1,'',bg=C['n_row']); ac(ws,row,2,'',bg=C['n_row'])
        ac(ws,row,3,'回答者数（ベース）',bold=True,bg=C['n_row'],fg=C['n_fg'],h='right')
        ac(ws,row,4,N,bold=True,bg=C['n_row'],fg=C['n_fg'],h='center')
        ac(ws,row,5,'',bg=C['n_row']); row += 1

        if qt == 'SA':
            res = tabulate_sa(qid,questions,data,hi)
            if not res: continue
            for oi,code in enumerate(res['codes']):
                label = res['items'].get(code,f'コード{code}')
                n_v = int(res['cnt'][code])
                pct = res['cnt'][code]/N*100 if N else 0
                bg = C['row_odd'] if oi%2==0 else C['row_even']
                ws.row_dimensions[row].height = 15
                ac(ws,row,1,qid,bg=bg,fg='888888',h='center')
                ac(ws,row,2,'',bg=bg)
                ac(ws,row,3,f'{code}. {label}',bg=bg)
                ac(ws,row,4,n_v,bg=bg,h='center')
                pc = ac(ws,row,5,fmt_pct(pct),bg=bg,h='center')
                if pct >= 50: pc.fill = hf(C['hi_green'])
                row += 1
        elif qt == 'MA':
            res = tabulate_ma(qid,questions,data,hi)
            if not res: continue
            for oi,item in enumerate(res['rows']):
                pct = item['n']/N*100 if N else 0
                bg = C['row_odd'] if oi%2==0 else C['row_even']
                ws.row_dimensions[row].height = 15
                ac(ws,row,1,qid,bg=bg,fg='888888',h='center')
                ac(ws,row,2,'',bg=bg)
                ac(ws,row,3,f'{item["no"]}. {item["text"]}',bg=bg)
                ac(ws,row,4,item['n'],bg=bg,h='center')
                pc = ac(ws,row,5,fmt_pct(pct),bg=bg,h='center')
                if pct >= 50: pc.fill = hf(C['hi_green'])
                row += 1
        row += 1

    wb.save(filepath)

# ============================================================
# Excel生成（クロス）
# ============================================================
def build_cross_excel(meta, questions, data, hi, cross_pairs, filepath):
    wb = Workbook(); wb.remove(wb.active)
    N = len(data); TODAY = datetime.date.today().strftime('%Y年%m月%d日')

    results = []
    for ax,tg in cross_pairs:
        res = cross_engine(ax,tg,questions,data,hi)
        if res: results.append((f'{ax}×{tg}',res))

    # --- 表紙 ---
    ws_toc = wb.create_sheet('表紙')
    ws_toc.sheet_view.showGridLines = False
    for col,w in zip('ABCDEF',[3,6,16,22,30,26]):
        ws_toc.column_dimensions[col].width = w
    ws_toc.merge_cells('B1:F1'); ws_toc.row_dimensions[1].height = 44
    t = ws_toc['B1']; t.value = 'クロス集計表'
    t.font = Font(bold=True,size=22,color='FFFFFF',name='Arial')
    t.fill = hf(C['h_navy']); t.alignment = al('center')
    infos = [('調査名',meta.get('調査名','—')),('調査ID',meta.get('調査ID','—')),
             ('有効回答数',f'{N:,} サンプル'),('出力日',TODAY)]
    for ri,(k,v) in enumerate(infos,3):
        ws_toc.row_dimensions[ri].height = 20
        kc = ws_toc.cell(ri,2,k)
        kc.font = Font(bold=True,color=C['h_fg']); kc.fill = hf(C['q_hd'])
        kc.alignment = al('center'); kc.border = thin(C['h_navy'])
        vc = ws_toc.cell(ri,3,v)
        vc.font = Font(size=9,name='Arial'); vc.fill = hf(C['gray'])
        vc.alignment = al(); vc.border = thin(C['h_navy'])
        ws_toc.merge_cells(start_row=ri,start_column=4,end_row=ri,end_column=6)
        ws_toc.cell(ri,4).border = thin(C['h_navy'])

    ws_toc.row_dimensions[8].height = 20
    for ci,lbl in enumerate(['No.','軸設問','集計設問','軸設問文','集計設問文'],2):
        c = ws_toc.cell(8,ci,lbl)
        c.font = Font(bold=True,color=C['h_fg'],size=9,name='Arial')
        c.fill = hf(C['h_navy']); c.alignment = al('center'); c.border = thin(C['h_navy'])

    toc_row = 9
    for no,(sname,res) in enumerate(results,1):
        bg = C['toc_odd'] if no%2==0 else C['toc_even']
        ws_toc.row_dimensions[toc_row].height = 17
        ac(ws_toc,toc_row,2,no,bg=bg,h='center')
        for ci,val in [(3,res['ax_qid']),(4,res['tg_qid'])]:
            lc = ws_toc.cell(toc_row,ci); lc.value = val
            safe = sname.replace("'","''")
            lc.hyperlink = f"#{safe}!A1"
            lc.font = Font(bold=True,color=C['link_fg'],underline='single',size=9,name='Arial')
            lc.fill = hf(bg); lc.alignment = al('center'); lc.border = thin()
        ac(ws_toc,toc_row,5,(res['ax_label'] or '')[:28],bg=bg)
        ac(ws_toc,toc_row,6,(res['tg_label'] or '')[:28],bg=bg)
        toc_row += 1

    # --- クロスシート ---
    for sname,res in results:
        ax_codes = res['ax_codes']; ax_items = res['ax_items']
        ax_bases = res['ax_bases']; rows_data = res['rows']
        n_ax = len(ax_codes); last_col = 3+n_ax

        ws = wb.create_sheet(sname[:31])
        ws.sheet_view.showGridLines = False
        ws.column_dimensions['A'].width = 34
        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 10
        for ci in range(4,4+n_ax):
            ws.column_dimensions[get_column_letter(ci)].width = 12

        ws.row_dimensions[1].height = 22
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=last_col-1)
        t = ws.cell(1,1,f'クロス集計：{res["ax_qid"]} × {res["tg_qid"]}')
        t.font = Font(bold=True,size=10,color='FFFFFF',name='Arial')
        t.fill = hf(C['h_navy']); t.alignment = al('center')
        bc = ws.cell(1,last_col,'▲表紙へ'); bc.hyperlink = '#表紙!A1'
        bc.font = Font(color=C['link_fg'],underline='single',size=8,italic=True,name='Arial')
        bc.fill = hf(C['back_bg']); bc.alignment = al('center'); bc.border = thin()

        for r_i,(label,txt) in enumerate([(f'【軸】{res["ax_label"]}',''),(f'【集計】{res["tg_label"]}','')],2):
            ws.row_dimensions[r_i].height = 24
            ws.merge_cells(start_row=r_i,start_column=1,end_row=r_i,end_column=last_col)
            cell = ws.cell(r_i,1,label)
            cell.font = Font(bold=True,size=9,color=C['q_fg'],name='Arial')
            cell.fill = hf(C['q_hd']); cell.alignment = al('left',wrap=True); cell.border = medium()

        ws.freeze_panes = 'A6'
        ws.row_dimensions[4].height = 22
        ac(ws,4,1,'選択肢',bold=True,bg=C['q_hd'],fg=C['h_fg'],h='center')
        ac(ws,4,2,'全体\nn',bold=True,bg=C['q_hd'],fg=C['h_fg'],h='center',wrap=True)
        ac(ws,4,3,'全体\n%',bold=True,bg=C['q_hd'],fg=C['h_fg'],h='center',wrap=True)
        for ci,ac_code in enumerate(ax_codes,4):
            ac(ws,4,ci,ax_items.get(ac_code,ac_code),bold=True,bg=C['q_hd'],fg=C['h_fg'],h='center',wrap=True)

        ws.row_dimensions[5].height = 16
        ac(ws,5,1,'回答者数（n）',bold=True,bg=C['n_row'],fg=C['n_fg'],h='right')
        ac(ws,5,2,res['N'],bold=True,bg=C['n_row'],fg=C['n_fg'],h='center')
        ac(ws,5,3,'',bg=C['n_row'])
        for ci,ac_code in enumerate(ax_codes,4):
            ac(ws,5,ci,ax_bases[ac_code],bold=True,bg=C['n_row'],fg=C['n_fg'],h='center')

        row = 6
        for oi,item in enumerate(rows_data):
            bg = C['row_odd'] if oi%2==0 else C['row_even']
            ws.row_dimensions[row].height = 15
            code_lbl = f'{item.get("code","") or item.get("no","")}. {item["label"]}'
            ac(ws,row,1,code_lbl,bg=bg)
            n_all = item.get('n_all',0)
            pct_all = n_all/res['N']*100 if res['N'] else 0
            ac(ws,row,2,n_all,bg=C['total_bg'],h='center',bold=True)
            pc = ac(ws,row,3,fmt_pct(pct_all),bg=C['total_bg'],h='center',bold=True)
            if pct_all >= 50: pc.fill = hf(C['hi_green'])
            for ci,ac_code in enumerate(ax_codes,4):
                by = item['by_ax'].get(ac_code,{})
                base = by.get('base',0); n_v = by.get('n',0)
                pct_v = n_v/base*100 if base else None
                if pct_v is None:
                    ac(ws,row,ci,'—',bg=bg,h='center',fg='AAAAAA')
                else:
                    cc = ac(ws,row,ci,fmt_pct(pct_v),bg=bg,h='center')
                    if pct_v >= 50: cc.fill = hf(C['hi_green'])
            row += 1

    wb.save(filepath)

# ============================================================
# APIエンドポイント
# ============================================================

@app.get("/")
def root():
    return {"status": "ok", "message": "for survey 集計API 稼働中"}

@app.get("/health")
def health():
    return {"status": "ok"}

@app.post("/api/tabulate")
async def tabulate(
    layout_file: UploadFile = File(...),
    rawdata_file: UploadFile = File(...),
    output_type: str = Form("both"),      # "gt" / "cross" / "both"
    cross_pairs: str = Form(""),          # "F1,Q1|F1,Q4|SC1,Q1" のようにパイプ区切り
):
    """
    レイアウト・ローデータをアップロードして集計Excelを返す
    """
    # 一時ディレクトリに保存
    tmp_dir = tempfile.mkdtemp()
    try:
        # ファイルをShift-JISで読み込む
        layout_bytes  = await layout_file.read()
        rawdata_bytes = await rawdata_file.read()

        layout_text  = layout_bytes.decode(ENCODING, errors='replace')
        rawdata_text = rawdata_bytes.decode(ENCODING, errors='replace')

        # パース
        meta, questions = parse_layout(layout_text)
        headers, hi, data = parse_rawdata(rawdata_text)

        if len(data) == 0:
            raise HTTPException(status_code=400,
                detail="有効回答（COMP）が見つかりません。ファイルを確認してください。")

        # 集計対象設問
        skip = {'MID','START','END','TIME','STA','GATE'}
        target_qids = [q for q in questions
                       if q not in skip and detect_type(questions[q]) in ('SA','MA')]

        # クロスペア
        pairs = []
        if cross_pairs.strip():
            for pair in cross_pairs.split('|'):
                parts = pair.split(',')
                if len(parts) == 2:
                    pairs.append((parts[0].strip(), parts[1].strip()))
        if not pairs:
            # デフォルトペア（先頭SA × 全SA/MA）
            sa_qs = [q for q in target_qids if detect_type(questions[q])=='SA'][:3]
            for ax in sa_qs[:1]:
                for tg in target_qids[:4]:
                    if ax != tg:
                        pairs.append((ax, tg))

        # Excel生成
        results = {}
        if output_type in ('gt','both'):
            gt_path = os.path.join(tmp_dir, 'GT.xlsx')
            build_gt_excel(meta, questions, data, hi, target_qids, gt_path)
            results['gt'] = gt_path

        if output_type in ('cross','both'):
            cross_path = os.path.join(tmp_dir, 'Cross.xlsx')
            build_cross_excel(meta, questions, data, hi, pairs, cross_path)
            results['cross'] = cross_path

        # 返却（両方の場合はGTを優先して返し、Crossは別途取得）
        # 簡易版：GTを返す（両方ならZIPにまとめる）
        if output_type == 'gt':
            return FileResponse(results['gt'],
                filename=f'{meta.get("調査ID","survey")}_GT.xlsx',
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        elif output_type == 'cross':
            return FileResponse(results['cross'],
                filename=f'{meta.get("調査ID","survey")}_Cross.xlsx',
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        else:
            # bothの場合はZIPにまとめる
            import zipfile
            zip_path = os.path.join(tmp_dir, 'report.zip')
            with zipfile.ZipFile(zip_path, 'w') as zf:
                zf.write(results['gt'],   f'{meta.get("調査ID","survey")}_GT.xlsx')
                zf.write(results['cross'], f'{meta.get("調査ID","survey")}_Cross.xlsx')
            return FileResponse(zip_path,
                filename=f'{meta.get("調査ID","survey")}_集計表.zip',
                media_type='application/zip')

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'集計エラー: {str(e)}')
    finally:
        # 一時ファイルは遅延削除（FileResponseが返った後）
        pass


@app.post("/api/check")
async def check_file(
    layout_file: UploadFile = File(...),
    rawdata_file: UploadFile = File(...),
):
    """ファイルの内容を事前確認する"""
    try:
        layout_bytes  = await layout_file.read()
        rawdata_bytes = await rawdata_file.read()
        layout_text  = layout_bytes.decode(ENCODING, errors='replace')
        rawdata_text = rawdata_bytes.decode(ENCODING, errors='replace')
        meta, questions = parse_layout(layout_text)
        headers, hi, data = parse_rawdata(rawdata_text)
        skip = {'MID','START','END','TIME','STA','GATE'}
        sa_qs = [q for q in questions if q not in skip and detect_type(questions[q])=='SA']
        ma_qs = [q for q in questions if q not in skip and detect_type(questions[q])=='MA']
        fa_qs = [q for q in questions if q not in skip and detect_type(questions[q])=='FA']
        return {
            "status"      : "ok",
            "survey_name" : meta.get('調査名','—'),
            "survey_id"   : meta.get('調査ID','—'),
            "valid_n"     : len(data),
            "sa_count"    : len(sa_qs),
            "ma_count"    : len(ma_qs),
            "fa_count"    : len(fa_qs),
            "sa_questions": sa_qs[:10],
            "ma_questions": ma_qs[:10],
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f'ファイルエラー: {str(e)}')


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
